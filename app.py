import streamlit as st
import pandas as pd
import io
import datetime
import zoneinfo
import os
import requests
import smtplib
import time
import hmac
import hashlib
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
# =============================================================================
# SECTION 1: SYSTEM CONFIGURATION & CONSTANTS
# =============================================================================
st.set_page_config(page_title="WOC - Time Tracking System", layout="wide", page_icon="📝")
# Endpoints for writing data
REGISTRATION_FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLSdY6ydD4YLYQEicFkk21DIRefUTT5ht8v4lbdZVr6hSbGOBAA/formResponse"
LOG_ENTRY_FORM_URL    = "https://docs.google.com/forms/d/1G8flLQrWJWGl5CwOEUe48zuAPre5mhJrbanx33uSkZk/formResponse"
# Endpoints for reading data
SHEET_ID       = "1zop4YKXKA1H8Iv89YwkGpP4c4YlGGFgz5jDYLT3psik"
TIMESHEETS_GID = "742432797"
ACCOUNTS_GID   = "1781560298"
# =============================================================================
# SECTION 2: CUSTOM CSS — color-coded rows, metric bar, mobile nav
# =============================================================================
st.markdown("""
<style>
/* ── Color-coded activity badge pills ── */
.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 12px;
    font-weight: 700;
    letter-spacing: 0.5px;
}
.badge-NOFA   { background:#DBEAFE; color:#1D4ED8; }
.badge-CARP   { background:#DCFCE7; color:#15803D; }
.badge-WOC    { background:#FEF3C7; color:#B45309; }
.badge-JJ     { background:#F3E8FF; color:#7E22CE; }
.badge-TRICAP { background:#FFE4E6; color:#BE123C; }
.badge-MDHHS  { background:#E0F2FE; color:#0369A1; }
.badge-BEWELL { background:#CCFBF1; color:#0F766E; }
.badge-OTHER  { background:#F1F5F9; color:#475569; }
/* ── Persistent hours bar ── */
.hours-bar {
    background: linear-gradient(135deg, #7B2CBF 0%, #9D4EDD 100%);
    border-radius: 12px;
    padding: 14px 24px;
    margin-bottom: 18px;
    display: flex;
    align-items: center;
    gap: 32px;
    flex-wrap: wrap;
}
.hours-bar .metric { text-align: center; }
.hours-bar .metric .val {
    font-size: 26px;
    font-weight: 800;
    color: #ffffff;
    line-height: 1;
}
.hours-bar .metric .lbl {
    font-size: 11px;
    color: #E0AAFF;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-top: 2px;
}
.hours-bar .divider {
    width: 1px; height: 40px;
    background: rgba(255,255,255,0.25);
}
.hours-bar .period-label {
    font-size: 13px;
    color: #E0AAFF;
    margin-left: auto;
}
/* ── Mobile-friendly nav ── */
@media (max-width: 640px) {
    .nav-btn { font-size: 12px !important; padding: 6px 4px !important; }
}
/* ── Admin panel highlight ── */
.admin-badge {
    display: inline-block;
    background: #7B2CBF;
    color: white;
    font-size: 11px;
    font-weight: 700;
    padding: 2px 10px;
    border-radius: 20px;
    letter-spacing: 1px;
    margin-left: 8px;
    vertical-align: middle;
}
/* ── Duplicate warning ── */
.dup-warning {
    background: #FEF9C3;
    border-left: 4px solid #EAB308;
    padding: 10px 16px;
    border-radius: 6px;
    font-size: 14px;
    margin-bottom: 8px;
}
</style>
""", unsafe_allow_html=True)
# =============================================================================
# SECTION 3: UI HEADERS & BADGE HELPERS
# =============================================================================
def render_woc_header():
    logo_filename = "woclogo.png"
    if os.path.exists(logo_filename):
        st.image(logo_filename, width=280)
    st.markdown(
        """
        <div style="background-color: #7B2CBF; padding: 15px 20px; border-radius: 10px;
                    margin-top: 10px; margin-bottom: 25px;">
            <h1 style="color: white; margin: 0; font-family: 'Calibri', sans-serif;
                       font-size: 30px; font-weight: bold; letter-spacing: 0.5px;">
                Women of Colors, Inc.</h1>
            <p style="color: #E0AAFF; margin: 4px 0 0 0; font-size: 16px;
                      font-family: 'Calibri', sans-serif;">
                Saginaw Community Prevention &amp; Training Program Hub</p>
        </div>
        """,
        unsafe_allow_html=True
    )
CODE_COLORS = {
    "NOFA":   "NOFA",
    "CARP":   "CARP",
    "WOC":    "WOC",
    "JJ":     "JJ",
    "TRICAP": "TRICAP",
    "MDHHS":  "MDHHS",
    "BEWELL": "BEWELL",
}
def code_badge(code):
    cls = CODE_COLORS.get(str(code).strip().upper(), "OTHER")
    return f'<span class="badge badge-{cls}">{code}</span>'
def hours_to_hhmm(decimal_hours):
    """Convert 1.75 → '1:45' safely with rounding to prevent 1:60 bugs"""
    try:
        total_minutes = int(round(float(decimal_hours) * 60))
        h = total_minutes // 60
        m = total_minutes % 60
        return f"{h}:{m:02d}"
    except Exception:
        return str(decimal_hours)
def safe_hours(value):
    """Convert a raw Hours cell (which may be blank, text, or numeric) to a float.

    Google Sheets sometimes returns blanks or stray text in numeric columns.
    The old astype(float) approach would either crash the whole page or silently
    corrupt totals. This always returns a clean float."""
    try:
        v = float(value)
        if v != v:  # NaN check
            return 0.0
        return v
    except (ValueError, TypeError):
        return 0.0
def safe_minutes(value):
    """Convert a raw Minutes cell to a clean int, tolerating blanks/NaN/text."""
    try:
        v = float(value)
        if v != v:  # NaN check
            return 0
        return int(round(v))
    except (ValueError, TypeError):
        return 0
# =============================================================================
# SECTION 4: DATA FETCHING & MAPPING HELPERS
# =============================================================================
def map_columns(df, rules):
    cols_map = {}
    for col in df.columns:
        c_lower = str(col).lower().strip()
        if ".1" in c_lower or ".2" in c_lower:
            continue
        for standard_name, keywords in rules.items():
            if standard_name not in cols_map:
                if any(kw in c_lower for kw in keywords):
                    cols_map[standard_name] = col
                    break
    if len(cols_map) >= max(1, len(rules) - 2):
        result = pd.DataFrame()
        for standard_name, actual_col in cols_map.items():
            result[standard_name] = df[actual_col]
        return result, True
    return df.copy(), False
def _csv_url(gid):
    return f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={gid}"
def _fetch_csv_with_retry(url, max_attempts=3):
    """Fetch a CSV URL with exponential backoff on failure."""
    for attempt in range(max_attempts):
        try:
            return pd.read_csv(url)
        except Exception as e:
            if attempt == max_attempts - 1:
                raise e
            time.sleep(2 ** attempt)
@st.cache_data(ttl=60)
def fetch_timesheets():
    try:
        raw = _fetch_csv_with_retry(_csv_url(TIMESHEETS_GID))
        rules = {
            "Timestamp":       ["timestamp"],
            "Date":            ["date"],
            "Instructor Name": ["instructor", "name", "staff"],
            "Time In":         ["time in"],
            "Time Out":        ["time out"],
            "Activity":        ["activity"],
            "Code":            ["code"],
            "Category":        ["category"],
            "Description":     ["description"],
            "Minutes":         ["minutes"],
            "Hours":           ["hours"],
        }
        df, mapped = map_columns(raw, rules)
        if not mapped:
            if len(raw.columns) >= 11:
                df = raw.copy()
                df.columns = (
                    ["Timestamp", "Date", "Instructor Name", "Time In", "Time Out",
                     "Activity", "Code", "Category", "Description", "Minutes", "Hours"]
                    + list(raw.columns[11:])
                )
                st.warning("⚠️ Timesheet columns could not be auto-detected — using positional "
                           "fallback. Check Google Sheet headers if data looks wrong.")
            else:
                st.error("🛑 Timesheet sheet structure is unrecognised. Contact your admin.")
                return pd.DataFrame(columns=["Timestamp","Date","Instructor Name","Time In",
                                             "Time Out","Activity","Code","Category",
                                             "Description","Minutes","Hours"])
        if "Code" in df.columns:
            df["Code"] = (df["Code"].astype(str).str.strip()
                            .replace({"MPHI": "CARP", "mphi": "CARP"}))
        # 🛠️ DATA HYGIENE: force Hours/Minutes to clean numbers ONCE at the source.
        # Blank cells, text, or stray characters in the Sheet previously crashed
        # astype(float)/astype(int) downstream or silently skewed totals.
        if "Hours" in df.columns:
            df["Hours"] = pd.to_numeric(df["Hours"], errors="coerce").fillna(0.0)
        if "Minutes" in df.columns:
            df["Minutes"] = pd.to_numeric(df["Minutes"], errors="coerce").fillna(0).astype(int)
        return df
    except Exception as e:
        st.error(f"🛑 Timesheet Database Connection Error (after retries): {e}")
        return pd.DataFrame(columns=["Timestamp","Date","Instructor Name","Time In",
                                     "Time Out","Activity","Code","Category",
                                     "Description","Minutes","Hours"])
@st.cache_data(ttl=60)
def fetch_accounts():
    try:
        raw = _fetch_csv_with_retry(_csv_url(ACCOUNTS_GID))
        rules = {
            "Timestamp":       ["timestamp"],
            "Instructor Name": ["instructor", "name"],
            "Email Address":   ["email"],
            "PIN":             ["pin"],
        }
        df, mapped = map_columns(raw, rules)
        if not mapped:
            if len(raw.columns) >= 4:
                df = raw.copy()
                df.columns = (["Timestamp", "Instructor Name", "Email Address", "PIN"]
                               + list(raw.columns[4:]))
                st.warning("⚠️ Accounts columns could not be auto-detected — using positional "
                           "fallback. Check Sheet headers if logins fail.")
            else:
                st.error("🛑 Accounts sheet structure is unrecognised. Contact your admin.")
                return pd.DataFrame(columns=["Timestamp","Instructor Name","Email Address","PIN"])
        return df
    except Exception as e:
        st.error(f"🛑 Account Profile Connection Error (after retries): {e}")
        return pd.DataFrame(columns=["Timestamp","Instructor Name","Email Address","PIN"])
# =============================================================================
# SECTION 5: EMAIL HELPER
# =============================================================================
def send_pin_email(recipient_email, recipient_name, user_pin):
    if "smtp" in st.secrets:
        try:
            APP_URL = "https://woc-time-tracker-ipumn9okebctvt3b3eqtj6.streamlit.app/"

            email_body = (
                f"Hello {recipient_name},\n\n"
                f"Your current PIN for the WOC Time Tracking Hub is: {user_pin}\n\n"
                f"HOW TO RESET YOUR PIN:\n"
                f"If you want to change your PIN to a new number, simply go to the app and use the 'Create Custom Account / PIN' tab. "
                f"Registering again with your exact same name and a new PIN will automatically overwrite this old one!\n\n"
                f"Access the app here: {APP_URL}\n\n"
                f"Regards,\nWomen of Colors Payroll Admin"
            )

            msg = MIMEText(email_body)
            msg['Subject'] = "WOC Time Tracker - PIN Recovery & Reset"
            msg['From']    = st.secrets["smtp"]["username"]
            msg['To']      = recipient_email

            with smtplib.SMTP_SSL(st.secrets["smtp"]["server"], int(st.secrets["smtp"]["port"])) as server:
                server.login(st.secrets["smtp"]["username"], st.secrets["smtp"]["password"])
                server.sendmail(st.secrets["smtp"]["username"], [recipient_email], msg.as_string())
            return True, "Success"
        except Exception as e:
            return False, str(e)
    return False, "Fallback"
def send_correction_email(instructor_name, entry_block, request_type, details):
    """Email an entry-correction request to the payroll admin.

    Recipient priority: st.secrets['admin']['correction_email'] if set,
    otherwise the SMTP username (the payroll admin's own account).
    Returns (ok, error_message)."""
    if "smtp" not in st.secrets:
        return False, "SMTP not configured"
    try:
        admin_addr = st.secrets.get("admin", {}).get(
            "correction_email", st.secrets["smtp"]["username"]
        )
        body = (
            f"A timesheet correction request was submitted from the WOC Time Tracking Hub.\n\n"
            f"REQUEST TYPE: {request_type}\n\n"
            f"ENTRY TO CORRECT:\n{entry_block}\n"
            f"REQUESTED CHANGE (in the instructor's words):\n{details if details.strip() else '(none provided — delete request)'}\n\n"
            f"HOW TO APPLY THIS FIX:\n"
            f"1. Open the Timesheets tab of the Google Sheet.\n"
            f"2. Find the row using the Sheet Timestamp shown above (or the date + times).\n"
            f"3. Edit the cells or delete the entire row.\n"
            f"4. The app refreshes its data within 60 seconds — no restart needed.\n\n"
            f"Regards,\nWOC Time Tracking Hub (automated message)"
        )
        msg = MIMEText(body)
        msg['Subject'] = f"Timesheet Correction Request — {instructor_name}"
        msg['From']    = st.secrets["smtp"]["username"]
        msg['To']      = admin_addr
        with smtplib.SMTP_SSL(st.secrets["smtp"]["server"], int(st.secrets["smtp"]["port"])) as server:
            server.login(st.secrets["smtp"]["username"], st.secrets["smtp"]["password"])
            server.sendmail(st.secrets["smtp"]["username"], [admin_addr], msg.as_string())
        return True, ""
    except Exception as e:
        return False, str(e)
# =============================================================================
# SECTION 5.5: SIGNED SESSION TOKENS  (fixes the ?user= login bypass)
# =============================================================================
# Previously, anyone could log in as any instructor by typing ?user=Name in
# the URL — the app trusted the name with no verification. Now the URL must
# also carry an expiry and an HMAC signature computed with a server-side
# secret. Without the secret, a valid signature cannot be forged, so typing
# ?user=Jane no longer works.
#
# Setup: add to your Streamlit secrets (any long random string, 30+ chars):
#   [auth]
#   token_secret = "paste-a-long-random-string-here"
#
# If token_secret is NOT configured, the app fails SAFE: no login URLs are
# written and sessions simply end on refresh (users log in again with PIN).
# The old insecure behavior is never used as a fallback.
#
# Note: a signed URL is still a "keep me logged in" link — anyone who copies
# someone's full URL is logged in as them until it expires. That's the same
# tradeoff as a browser cookie; tokens expire after TOKEN_DAYS_VALID days.
TOKEN_DAYS_VALID = 14
def _auth_secret():
    return str(st.secrets.get("auth", {}).get("token_secret", "")).strip()
def make_login_token(name, is_admin):
    """Create (expiry, signature) for a login URL, or None if no secret set."""
    secret = _auth_secret()
    if not secret:
        return None
    exp     = int(time.time()) + TOKEN_DAYS_VALID * 86400
    role    = "admin" if is_admin else "user"
    payload = f"{name.strip().lower()}|{exp}|{role}"
    sig     = hmac.new(secret.encode(), payload.encode(), hashlib.sha256).hexdigest()
    return exp, sig
def verify_login_token(name, exp_str, role, sig):
    """Return True only if the URL's signature is valid and unexpired."""
    secret = _auth_secret()
    if not secret or not name or not exp_str or not sig or role not in ("admin", "user"):
        return False
    try:
        exp = int(exp_str)
    except (ValueError, TypeError):
        return False
    if time.time() > exp:
        return False
    payload  = f"{str(name).strip().lower()}|{exp}|{role}"
    expected = hmac.new(secret.encode(), payload.encode(), hashlib.sha256).hexdigest()
    return hmac.compare_digest(expected, str(sig))
def set_login_query_params(name, is_admin):
    """Write a signed stay-logged-in URL, or none at all if no secret is set."""
    token = make_login_token(name, is_admin)
    if token:
        exp, sig = token
        st.query_params["user"] = name
        st.query_params["exp"]  = str(exp)
        st.query_params["role"] = "admin" if is_admin else "user"
        st.query_params["auth"] = sig
    else:
        # Fail safe: never write an unauthenticated ?user= URL.
        st.query_params.clear()
# =============================================================================
# SECTION 6: DUPLICATE DETECTION HELPER
# =============================================================================
def check_duplicate(entry_date, time_in, time_out, name, df):
    """Returns True if an identical entry already exists in the cached data."""
    if df.empty or "ParsedDate" not in df.columns:
        return False
    match = df[
        (df["Instructor Name"].astype(str).str.strip().str.lower() == name.strip().lower()) &
        (df["ParsedDate"] == entry_date) &
        (df["Time In"].astype(str).str.strip()  == time_in.strip()) &
        (df["Time Out"].astype(str).str.strip() == time_out.strip())
    ]
    return not match.empty
def check_overlap(entry_date, start_dt, end_dt, name, df):
    """Returns a list of existing same-day entries whose time range overlaps
    the proposed one. Catches partial double-counting the exact-duplicate
    check misses (e.g. logging 9:00–10:00 and then 9:30–10:30).

    Back-to-back entries (9:00–10:00 followed by 10:00–11:00) do NOT count
    as overlapping. Rows with unparseable times are skipped rather than
    blocking the save."""
    if df.empty or "ParsedDate" not in df.columns:
        return []
    day_rows = df[
        (df["Instructor Name"].astype(str).str.strip().str.lower() == name.strip().lower()) &
        (df["ParsedDate"] == entry_date)
    ]
    conflicts = []
    for _, row in day_rows.iterrows():
        try:
            ex_start = datetime.datetime.strptime(
                f"{entry_date} {str(row.get('Time In','')).strip()}", "%Y-%m-%d %I:%M %p")
            ex_end   = datetime.datetime.strptime(
                f"{entry_date} {str(row.get('Time Out','')).strip()}", "%Y-%m-%d %I:%M %p")
        except (ValueError, TypeError):
            continue
        # Standard interval-overlap test with strict inequalities so that
        # touching boundaries (10:00 end / 10:00 start) are allowed.
        if start_dt < ex_end and end_dt > ex_start:
            conflicts.append(
                f"{str(row.get('Time In','')).strip()} – {str(row.get('Time Out','')).strip()}"
                f" ({str(row.get('Activity','')).strip()})"
            )
    return conflicts
# =============================================================================
# SECTION 7: EXCEL GENERATION HELPERS (Hoisted for Performance)
# =============================================================================
def _make_styles():
    """Shared Excel Style Factory"""
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    shaded   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    subtotal = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    return {
        "title":    Font(name="Calibri", size=14, bold=True),
        "bold":     Font(name="Calibri", size=11, bold=True),
        "regular":  Font(name="Calibri", size=11),
        "small":    Font(name="Calibri", size=9, italic=True),
        "cursive":  Font(name="Brush Script MT", size=16, italic=True, color="002060"),
        "check_ok": Font(name="Calibri", size=11, bold=True, color="15803D"),
        "thin":     thin,
        "shaded":   shaded,
        "subtotal": subtotal,
    }
# Instructors whose exported timesheet should show Vicki Hill as manager.
# Names are compared lowercased + stripped, so capitalization or stray spaces
# in how the name is stored won't break the match. Add name variants here if
# either person is ever entered with a middle initial, etc.
SPECIAL_MANAGER_STAFF = {"jawain swint", "evelyn mcgovern"}
DEFAULT_MANAGER_NAME  = "Evelyn McGovern"
SPECIAL_MANAGER_NAME  = "Vicki Hill"
@st.cache_data(ttl=60)
def build_timesheet_bytes(instr, p_start, p_end, period_data_json, today_str):
    """Cached Excel generation — only rebuilds when data or period changes.

    ACCURACY FIXES (v2):
    1. All hour values written to cells are rounded to 2 decimals. Repeated
       float addition previously produced cells like 2.4899999999999998,
       which made totals look wrong in Excel.
    2. Hours are coerced with pd.to_numeric — blank/text cells no longer
       crash generation or silently drop from totals.
    3. A 'Column Subtotals' row now sums each grant column AND the Hours
       Worked column using live Excel SUM formulas, so the printed subtotals
       can never disagree with the visible daily cells.
    4. A 'Balance Check' cell verifies (in Excel, live) that the grant
       columns add up to the Hours Worked total, flagging any mismatch.
    5. Hours logged under an unrecognized grant code are no longer silently
       dropped from the grant columns — they are totaled and flagged in a
       warning row so payroll can see exactly where a mismatch comes from.
    6. 'Actual Hours Worked' is now a formula pointing at the subtotal row
       instead of a separately computed number, guaranteeing agreement.

    MANAGER OVERRIDE:
    Jawain Swint and Evelyn McGovern report to Vicki Hill; everyone else
    reports to Evelyn McGovern. Because `instr` is part of the cache key,
    the two manager versions cache separately with no collision.
    """
    period_data = pd.read_json(io.StringIO(period_data_json), orient="records")
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Sheet"
    ws.sheet_view.showGridLines  = True
    ws.print_options.gridLines   = True
    ws.page_setup.orientation    = 'landscape'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    S = _make_styles()
    # Manager override: Jawain Swint and Evelyn McGovern report to Vicki Hill;
    # everyone else reports to Evelyn McGovern.
    manager_name = (SPECIAL_MANAGER_NAME
                    if str(instr).strip().lower() in SPECIAL_MANAGER_STAFF
                    else DEFAULT_MANAGER_NAME)
    ws["A1"] = "BiWeekly Employee Time Sheet";  ws["A1"].font = S["title"]
    ws["A2"] = "Women of Colors";               ws["A2"].font = S["bold"]
    ws["A3"] = "Employee Details:";             ws["A3"].font = S["bold"]
    ws["D3"] = f"Name :  {instr}";              ws["D3"].font = S["regular"]
    ws["F3"] = "Email: payroll@yeoandyeo.com";  ws["F3"].font = S["regular"]
    ws["A4"] = "Manager Details:";              ws["A4"].font = S["bold"]
    ws["D4"] = f"Name: {manager_name}";         ws["D4"].font = S["regular"]
    ws["F4"] = "Fax: 989-793-0186";             ws["F4"].font = S["regular"]
    ws["A5"] = f"Period Start Date: {p_start.strftime('%m/%d/%Y')}"; ws["A5"].font = S["bold"]
    ws["E5"] = f"Period End Date:  {p_end.strftime('%m/%d/%Y')}";    ws["E5"].font = S["bold"]
    # Grant code columns — extended to column 13 to include BeWell
    for col_idx, text in enumerate(["","","","","","Total Hours","NOFA","WOC","JJ","TRICAP","CARP","MDHHS","BeWell"], 1):
        if text:
            c = ws.cell(row=7, column=col_idx, value=text)
            c.font = S["bold"]
            c.alignment = Alignment(horizontal="center", wrap_text=True)
    for col_idx, text in enumerate(["","Day","Date","Time In","Time Out","Hours Worked","NOFA","WOC","JJ","TRICAP","CARP","MDHHS","BeWell"], 1):
        if text:
            c = ws.cell(row=9, column=col_idx, value=text)
            c.font = S["bold"]
            c.alignment = Alignment(horizontal="center")
    code_col_map   = {"NOFA": 7, "WOC": 8, "JJ": 9, "TRICAP": 10, "CARP": 11, "MDHHS": 12, "BEWELL": 13}
    days_in_period = max(1, (p_end - p_start).days + 1)
    date_list      = [p_start + datetime.timedelta(days=x) for x in range(days_in_period)]
    if "Date" in period_data.columns:
        period_data["ParsedDate"] = pd.to_datetime(period_data["Date"], errors='coerce').dt.date
    # 🛠️ Numeric hygiene: coerce Hours ONCE so every sum below is trustworthy.
    if "Hours" in period_data.columns:
        period_data["Hours"] = pd.to_numeric(period_data["Hours"], errors="coerce").fillna(0.0)
    r = 10
    first_data_row  = r
    unmapped_hours  = 0.0
    unmapped_codes  = set()
    for idx, d in enumerate(date_list):
        if idx == 7:
            r += 1
        ws.cell(row=r, column=2, value=d.strftime("%A")).font       = S["regular"]
        ws.cell(row=r, column=3, value=d.strftime("%Y-%m-%d")).font = S["regular"]
        day_logs = period_data[period_data['ParsedDate'] == d] if "ParsedDate" in period_data.columns else pd.DataFrame()
        if not day_logs.empty:
            ws.cell(row=r, column=4, value=" / ".join(day_logs['Time In'].astype(str).tolist())).font  = S["regular"]
            ws.cell(row=r, column=5, value=" / ".join(day_logs['Time Out'].astype(str).tolist())).font = S["regular"]
            # 🛠️ Round the daily total so the cell holds a clean 2-decimal value
            ws.cell(row=r, column=6, value=round(float(day_logs['Hours'].sum()), 2)).font = S["regular"]

            for c_idx in range(7, 14):
                c = ws.cell(row=r, column=c_idx, value=0)
                c.font = S["regular"]; c.fill = S["shaded"]; c.border = S["thin"]
            for _, rl in day_logs.iterrows():
                code = str(rl.get('Code', '')).strip().upper()
                hw   = safe_hours(rl.get('Hours', 0.0))
                if code in code_col_map:
                    ci  = code_col_map[code]
                    cv  = ws.cell(row=r, column=ci).value or 0.0
                    # 🛠️ Round after every accumulation to kill float drift
                    ac  = ws.cell(row=r, column=ci, value=round(float(cv) + hw, 2))
                    ac.font = S["regular"]; ac.border = S["thin"]
                    ac.fill = PatternFill(fill_type=None)
                elif hw > 0:
                    # 🛠️ Previously these hours vanished from the grant columns
                    # with no trace — the #1 cause of "totals don't add up".
                    unmapped_hours += hw
                    unmapped_codes.add(code if code else "(blank)")
        else:
            for c_idx in range(4, 14):
                c = ws.cell(row=r, column=c_idx, value=0)
                c.font = S["regular"]; c.fill = S["shaded"]; c.border = S["thin"]
        r += 1
    last_data_row = r - 1
    # ── NEW: COLUMN SUBTOTALS ROW (live Excel SUM formulas) ──────────────────
    subtotal_row = r
    lbl = ws.cell(row=subtotal_row, column=2, value="Column Subtotals:")
    lbl.font = S["bold"]; lbl.fill = S["subtotal"]
    ws.cell(row=subtotal_row, column=3).fill = S["subtotal"]
    ws.cell(row=subtotal_row, column=4).fill = S["subtotal"]
    ws.cell(row=subtotal_row, column=5).fill = S["subtotal"]
    for c_idx in range(6, 14):
        col_letter = get_column_letter(c_idx)
        cell = ws.cell(
            row=subtotal_row, column=c_idx,
            value=f"=ROUND(SUM({col_letter}{first_data_row}:{col_letter}{last_data_row}),2)"
        )
        cell.font = S["bold"]; cell.fill = S["subtotal"]; cell.border = S["thin"]
        cell.number_format = "0.00"
    r += 1
    # ── NEW: BALANCE / VALIDATION CHECK ROW ──────────────────────────────────
    # Verifies inside Excel itself that grant columns (G:M) sum to the Hours
    # Worked column (F). If anyone edits the file later, the check stays live.
    check_row = r
    ws.cell(row=check_row, column=2, value="Grant Columns Total:").font = S["bold"]
    gtot = ws.cell(row=check_row, column=3,
                   value=f"=ROUND(SUM(G{subtotal_row}:M{subtotal_row}),2)")
    gtot.font = S["bold"]; gtot.number_format = "0.00"
    ws.cell(row=check_row, column=5, value="Balance Check:").font = S["bold"]
    chk = ws.cell(
        row=check_row, column=6,
        value=(
            f'=IF(ABS(F{subtotal_row}-SUM(G{subtotal_row}:M{subtotal_row}))<0.02,'
            f'"BALANCED - OK",'
            f'"MISMATCH: "&TEXT(F{subtotal_row}-SUM(G{subtotal_row}:M{subtotal_row}),"0.00")&" hrs unaccounted")'
        )
    )
    chk.font = S["check_ok"]
    r += 1
    # ── NEW: UNMAPPED-CODE WARNING ROW (only appears when needed) ────────────
    if unmapped_hours > 0.009:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
        warn = ws.cell(
            row=r, column=2,
            value=(f"WARNING: {round(unmapped_hours, 2):.2f} hrs were logged under "
                   f"unrecognized grant code(s) [{', '.join(sorted(unmapped_codes))}] and are "
                   f"included in Hours Worked but NOT in any grant column. Correct these "
                   f"entries in the log to balance this timesheet.")
        )
        warn.font = Font(name="Calibri", size=10, bold=True, color="B45309")
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="Total Target Hours:").font  = S["bold"]
    ws.cell(row=r, column=3, value=75.0).font                   = S["bold"]
    r += 1
    ws.cell(row=r, column=2, value="Actual Hours Worked:").font = S["bold"]
    # 🛠️ These now reference the subtotal formula, so they can never disagree
    # with the daily grid above (the old version computed them separately).
    actual_c3 = ws.cell(row=r, column=3, value=f"=F{subtotal_row}")
    actual_c3.font = S["bold"]; actual_c3.number_format = "0.00"
    actual_c6 = ws.cell(row=r, column=6, value=f"=F{subtotal_row}")
    actual_c6.font = S["bold"]; actual_c6.number_format = "0.00"
    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
    cert = ws.cell(row=r, column=2, value="CLIENT: I CERTIFY THAT THE HOURS WORKED ON THIS TIME SLIP ARE CORRECT.")
    cert.font = S["bold"]
    cert.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 24
    r += 2
    ws.cell(row=r, column=2, value="Employee Signature:").font = S["bold"]
    ws.cell(row=r, column=3, value=instr).font                 = S["cursive"]
    ws.cell(row=r, column=5, value="Date:").font               = S["bold"]
    ws.cell(row=r, column=6, value=today_str).font             = S["regular"]
    r += 2
    ws.cell(row=r, column=2, value="Manager Signature:").font  = S["bold"]
    ws.cell(row=r, column=5, value="Date:").font               = S["bold"]
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row in [1,2,3,4,5] or "CLIENT: I CERTIFY" in val_str or "WARNING:" in val_str or val_str.startswith("="):
                continue
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
@st.cache_data(ttl=60)
def build_additional_bytes(instr, p_start, p_end, period_data_json):
    """Cached additional hours report generation."""
    period_data = pd.read_json(io.StringIO(period_data_json), orient="records")
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Form"
    ws.sheet_view.showGridLines  = True
    ws.print_options.gridLines   = True
    ws.page_setup.orientation    = 'landscape'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    S = _make_styles()
    ws["A1"] = f"Additional Hours Report FY20 - Due {p_start.strftime('%m/%d/%y')} - {p_end.strftime('%m/%d/%Y')}"
    ws["A1"].font = S["bold"]
    ws["A2"] = "Agency Name";              ws["A2"].font = S["bold"]
    ws["C2"] = "ADDITIONAL HOURS REPORT";  ws["C2"].font = S["bold"]
    for col_idx, h in enumerate(["Date","Staff Name","Category","Description","Time in minutes"], 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = S["bold"]
        c.alignment = Alignment(horizontal="left")
    curr_row   = 4
    total_mins = 0
    for _, log in period_data.iterrows():
        # 🛠️ safe_minutes tolerates blank/NaN/text cells that previously
        # crashed report generation via int() or corrupted the total.
        mins = safe_minutes(log.get('Minutes', 0))
        ws.cell(row=curr_row, column=1, value=str(log.get('Date',''))).font        = S["regular"]
        ws.cell(row=curr_row, column=2, value=instr).font                          = S["regular"]
        ws.cell(row=curr_row, column=3, value=str(log.get('Category',''))).font    = S["regular"]
        ws.cell(row=curr_row, column=4, value=str(log.get('Description',''))).font = S["regular"]
        ws.cell(row=curr_row, column=5, value=mins).font                           = S["regular"]
        total_mins += mins
        curr_row   += 1
    ws.cell(row=curr_row, column=4, value="Total").font    = S["bold"]
    # 🛠️ Live Excel formula so the total always matches the listed rows
    total_cell = ws.cell(row=curr_row, column=5,
                         value=f"=SUM(E4:E{max(4, curr_row - 1)})" if curr_row > 4 else 0)
    total_cell.font = S["bold"]
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
# =============================================================================
# SECTION 8: LOGIN / REGISTRATION PORTAL
# =============================================================================
existing_data    = fetch_timesheets()
account_registry = fetch_accounts()
if "user" in st.query_params and "logged_in" not in st.session_state:
    # ── SECURE AUTO-LOGIN: only honor the URL if its signature verifies ──
    _qp_user = st.query_params.get("user", "")
    _qp_exp  = st.query_params.get("exp",  "")
    _qp_role = st.query_params.get("role", "user")
    _qp_sig  = st.query_params.get("auth", "")
    if verify_login_token(_qp_user, _qp_exp, _qp_role, _qp_sig):
        st.session_state["logged_in"]       = True
        st.session_state["instructor_name"] = _qp_user
        st.session_state["is_admin"]        = (_qp_role == "admin")
    else:
        # Forged, tampered, or expired link — clear it and show the login page.
        st.query_params.clear()
if not st.session_state.get("logged_in"):
    render_woc_header()
    st.markdown("<h3 style='color: #7B2CBF; margin-bottom: 10px;'>🔐 Instructor Access Hub</h3>",
                unsafe_allow_html=True)
    portal_tab = st.radio(
        "Choose Action:",
        ["Sign In", "Create Custom Account / PIN", "Forgot PIN / Reset Option"],
        horizontal=True, label_visibility="collapsed"
    )
    col_portal, _ = st.columns([1.5, 2])
    with col_portal:
        # ── SIGN IN ──────────────────────────────────────────────────────────
        if portal_tab == "Sign In":
            with st.form("signin_panel"):
                login_name = st.text_input("Instructor Name:", placeholder="First & Last Name")
                login_pin  = st.text_input("Enter Personal PIN:", type="password",
                                             placeholder="Type your PIN")
                submit_login = st.form_submit_button("🔓 Log In")
                if submit_login:
                    cleaned_name  = login_name.strip()
                    matched_users = account_registry[
                        account_registry["Instructor Name"]
                            .astype(str).str.strip().str.lower() == cleaned_name.lower()
                    ]
                    if not matched_users.empty:
                        # 🛠️ SMART PIN CHECK: Safely handles stripped zeros from Google Sheets
                        raw_correct_pin = str(matched_users.iloc[-1]["PIN"]).strip()

                        # Clean up pandas float conversion if it happened
                        if raw_correct_pin.endswith('.0'):
                            raw_correct_pin = raw_correct_pin[:-2]

                        entered_pin = login_pin.strip()
                        is_match = False

                        # 1. First check exact string match
                        if entered_pin == raw_correct_pin:
                            is_match = True

                        # 2. Second check mathematical match (allows "0528" to match "528")
                        if not is_match:
                            try:
                                if int(entered_pin) == int(float(raw_correct_pin)):
                                    is_match = True
                            except ValueError:
                                pass

                        # 3. Third check string zero-strip match (fallback)
                        if not is_match:
                            if entered_pin.lstrip('0') == raw_correct_pin.lstrip('0') and len(entered_pin.lstrip('0')) > 0:
                                is_match = True
                        if is_match:
                            st.session_state["logged_in"]       = True
                            st.session_state["instructor_name"] = cleaned_name
                            st.session_state["is_admin"]        = False
                            set_login_query_params(cleaned_name, is_admin=False)
                            st.rerun()
                        else:
                            st.error("Authentication Error: Invalid PIN for this profile.")
                    else:
                        master_pw = st.secrets.get("admin", {}).get("master_password", "")
                        if master_pw and login_pin.strip() == master_pw:
                            st.session_state["logged_in"]       = True
                            st.session_state["instructor_name"] = cleaned_name
                            st.session_state["is_admin"]        = True
                            set_login_query_params(cleaned_name, is_admin=True)
                            st.rerun()
                        else:
                            st.error(f"Profile Not Found: '{cleaned_name}' is not registered yet.")
        # ── REGISTRATION ─────────────────────────────────────────────────────
        elif portal_tab == "Create Custom Account / PIN":
            st.info("💡 Setting up your profile connects your name to a custom passcode.")
            with st.form("registration_panel"):
                reg_name  = st.text_input("Full Instructor Name:", placeholder="First and Last Name")
                reg_email = st.text_input("Email Address:",        placeholder="username@domain.com")
                reg_pin   = st.text_input("Create 4-to-6 Digit PIN:", type="password",
                                          placeholder="Choose your passcode")
                submit_reg = st.form_submit_button("📝 Register Account")
                if submit_reg:
                    if not reg_name.strip() or not reg_email.strip() or not reg_pin.strip():
                        st.error("Validation Error: All fields are required.")
                    else:
                        acc_data = {
                            "entry.576544689":  reg_name.strip(),
                            "entry.836662014":  reg_email.strip(),
                            "entry.2099667226": reg_pin.strip()
                        }
                        with st.spinner("Creating your secure profile..."):
                            try:
                                res = requests.post(REGISTRATION_FORM_URL, data=acc_data, timeout=10)
                                if res.ok:
                                    # 🛠️ CACHE CLEAR: Forces the app to download the new PIN instantly
                                    fetch_accounts.clear()
                                    st.success("Account created! Switch to 'Sign In' to enter.")
                                else:
                                    st.error(f"Database Error (Code {res.status_code}).")
                            except Exception as e:
                                st.error(f"Network Connection Failed: {e}")
        # ── PIN RECOVERY ──────────────────────────────────────────────────────
        elif portal_tab == "Forgot PIN / Reset Option":
            with st.form("recovery_panel"):
                recover_email   = st.text_input("Registered Email:", placeholder="username@domain.com")
                submit_recovery = st.form_submit_button("🔍 Retrieve Access Passcode")
                if submit_recovery:
                    matched_emails = account_registry[
                        account_registry["Email Address"].astype(str).str.strip().str.lower()
                        == recover_email.strip().lower()
                    ]
                    if not matched_emails.empty:
                        user_account = matched_emails.iloc[-1]
                        found_name   = user_account["Instructor Name"]

                        # Clean up PIN formatting for the email
                        found_pin = str(user_account["PIN"]).strip()
                        if found_pin.endswith('.0'):
                            found_pin = found_pin[:-2]

                        with st.spinner("Dispatching secure recovery instructions..."):
                            status, _ = send_pin_email(recover_email.strip(), found_name, found_pin)

                        if status:
                            st.success(f"📬 Recovery instructions sent to {recover_email.strip()}.")
                        else:
                            st.error("⚠️ The automated email system is currently offline. Please contact the Payroll Administrator directly to recover your PIN.")
                    else:
                        st.error("That email is not in our registry.")
    st.stop()
# =============================================================================
# SECTION 9: TIMEZONE-AWARE TODAY  (America/Detroit = Saginaw, MI)
# =============================================================================
TODAY = datetime.datetime.now(zoneinfo.ZoneInfo("America/Detroit")).date()
# =============================================================================
# SECTION 10: AUTHENTICATED SESSION HEADER
# =============================================================================
render_woc_header()
instructor_input = st.session_state["instructor_name"]
is_admin         = st.session_state.get("is_admin", False)
# Trigger Success Toast after a page refresh
if "toast_message" in st.session_state:
    st.toast(st.session_state["toast_message"], icon="🎉")
    del st.session_state["toast_message"]
col_user1, col_user2 = st.columns([3, 1])
with col_user1:
    admin_badge = '<span class="admin-badge">ADMIN</span>' if is_admin else ""
    st.markdown(f"#### Welcome back, **{instructor_input}**! 👋 {admin_badge}", unsafe_allow_html=True)
with col_user2:
    if st.button("🚪 Log Out / Clear Session", use_container_width=True):
        st.session_state.clear()
        st.query_params.clear()
        st.rerun()
# =============================================================================
# SECTION 11: PAY PERIOD NAVIGATION
# =============================================================================
if "period_offset" not in st.session_state:
    st.session_state.period_offset = 0
ANCHOR_DATE       = datetime.date(2026, 5, 24)
days_since_anchor = (TODAY - ANCHOR_DATE).days
completed_periods = days_since_anchor // 14
active_period_index = completed_periods + st.session_state.period_offset
auto_period_start   = ANCHOR_DATE + datetime.timedelta(days=active_period_index * 14)
auto_period_end     = auto_period_start + datetime.timedelta(days=13)
st.subheader("🗓️ Pay Period Review Settings")
col_nav1, col_nav2, col_nav3 = st.columns([1, 2, 1])
with col_nav1:
    if st.button("⬅️ Previous", use_container_width=True, key="nav_prev"):
        st.session_state.period_offset -= 1
        st.rerun()
with col_nav2:
    col_date, col_reset = st.columns([2, 1])
    with col_date:
        st.markdown(
            f"<h4 style='text-align:center;margin-top:5px;color:#7B2CBF;'>"
            f"{auto_period_start.strftime('%b %d')} — {auto_period_end.strftime('%b %d, %Y')}"
            f"</h4>", unsafe_allow_html=True
        )
    with col_reset:
        if st.session_state.period_offset != 0:
            if st.button("🔄 Current", use_container_width=True, help="Jump back to the active pay period"):
                st.session_state.period_offset = 0
                st.rerun()
with col_nav3:
    if st.button("Next ➡️", use_container_width=True, key="nav_next"):
        st.session_state.period_offset += 1
        st.rerun()
col_p1, col_p2 = st.columns(2)
with col_p1:
    pay_period_start = st.date_input("Start Date (editable override)", value=auto_period_start)
with col_p2:
    pay_period_end   = st.date_input("End Date (editable override)",   value=auto_period_end)
if pay_period_start != auto_period_start or pay_period_end != auto_period_end:
    st.info("ℹ️ Manual date override active. Use nav buttons to return to automatic mode.")
st.markdown("---")
# =============================================================================
# SECTION 12: FILTER DATA
# =============================================================================
total_database_records = 0
current_period_df      = pd.DataFrame()
running_hours          = 0.0
running_minutes        = 0
all_instructors_df     = pd.DataFrame()
last_time_in_str       = None   # user's most recently logged Time In (for smart defaults)
last_time_out_str      = None   # user's most recently logged Time Out
if not existing_data.empty and "Instructor Name" in existing_data.columns:
    all_instructors_df = existing_data.copy()
    if "Date" in all_instructors_df.columns:
        all_instructors_df["ParsedDate"] = pd.to_datetime(all_instructors_df["Date"], errors='coerce').dt.date
        all_instructors_df = all_instructors_df.dropna(subset=["ParsedDate"])
    if instructor_input.strip():
        user_filtered_df = all_instructors_df[
            all_instructors_df["Instructor Name"].astype(str).str.strip().str.lower() == instructor_input.strip().lower()
        ].copy()
        if not user_filtered_df.empty:
            # ── SMART TIME DEFAULTS: remember the user's most recent entry ──
            # Sorted by the Sheet's Timestamp when parseable; otherwise sheet
            # order (newest rows last) is already a good proxy for recency.
            _recent = user_filtered_df.copy()
            if "Timestamp" in _recent.columns:
                _recent["_ts"] = pd.to_datetime(_recent["Timestamp"], errors="coerce")
                _recent = _recent.sort_values("_ts", na_position="first")
            _last_row = _recent.iloc[-1]
            last_time_in_str  = str(_last_row.get("Time In", "")).strip()
            last_time_out_str = str(_last_row.get("Time Out", "")).strip()
            current_period_df = user_filtered_df[
                (user_filtered_df["ParsedDate"] >= pay_period_start) &
                (user_filtered_df["ParsedDate"] <= pay_period_end)
            ].sort_values(by="ParsedDate", ascending=True)
            total_database_records = len(user_filtered_df)
            # 🛠️ pd.to_numeric replaces astype(float)/astype(int), which crashed
            # the whole page whenever the Sheet contained a blank or text cell.
            if 'Hours' in current_period_df.columns:
                running_hours = round(float(pd.to_numeric(current_period_df['Hours'], errors='coerce').fillna(0).sum()), 2)
            if 'Minutes' in current_period_df.columns:
                running_minutes = int(pd.to_numeric(current_period_df['Minutes'], errors='coerce').fillna(0).sum())
# =============================================================================
# SECTION 13: PERSISTENT HOURS BAR
# =============================================================================
hhmm = hours_to_hhmm(running_hours)
period_label = f"{pay_period_start.strftime('%b %d')} – {pay_period_end.strftime('%b %d, %Y')}"
st.markdown(f"""
<div class="hours-bar">
    <div class="metric">
        <div class="val">{running_hours:.2f}</div>
        <div class="lbl">Hours This Period</div>
    </div>
    <div class="divider"></div>
    <div class="metric">
        <div class="val">{hhmm}</div>
        <div class="lbl">HH:MM Format</div>
    </div>
    <div class="divider"></div>
    <div class="metric">
        <div class="val">{running_minutes}</div>
        <div class="lbl">Total Minutes</div>
    </div>
    <div class="divider"></div>
    <div class="metric">
        <div class="val">{len(current_period_df)}</div>
        <div class="lbl">Entries Logged</div>
    </div>
    <div class="period-label">📅 {period_label}</div>
</div>
""", unsafe_allow_html=True)
# =============================================================================
# SECTION 14: ACTIVITY STRUCTURE — Activity + Grant/Program two-step picker
# =============================================================================
# Each activity lists the grant/program choices staff can bill it to.
# Single-choice activities auto-lock the second dropdown, so nothing can be
# miscoded. The saved Activity string stays "Activity - Grant" for multi-grant
# activities and just the activity name for single-grant ones, keeping old
# sheet history consistent.
ACTIVITY_STRUCTURE = {
    "Botvin Life Skills Training":         ["NOFA"],
    "Botvin/Tricap":                       ["Tri-Cap"],
    "Carp / MHEF":                         ["CARP"],
    "Office Admin":                        ["NOFA", "CARP", "MDHHS"],
    "PFL (Training/Data Entry)":           ["NOFA"],
    "PFL Instructor Training (Juvenile)":  ["JJ"],
    "PFL Instructor Training (Tri-Cap)":   ["Tri-Cap"],
    "Pathway To Purpose":                  ["JJ"],
    "Prevention Team Meeting":             ["NOFA"],
    "Sick":                                ["NOFA", "CARP", "JJ", "MDHHS", "BeWell"],
    "Training":                            ["Botvin", "PFL", "NOFA", "CARP", "JJ", "MDHHS", "BeWell"],
    "WOC Facility Maintenance":            ["WOC"],
    "WOC IT Support":                      ["WOC"],
}
# Maps every possible grant/program choice to its official funding code
GRANT_TO_CODE = {
    "NOFA":        "NOFA",
    "HF or CARP":  "CARP",
    "JJ":          "JJ",
    "MDHHS":       "MDHHS",
    "BeWell":      "BEWELL",
    "Botvin":      "NOFA",   # Training - Botvin bills to NOFA
    "PFL":         "NOFA",   # Training - PFL bills to NOFA
    "Tri-Cap":     "TRICAP",
    "WOC":         "WOC",
}
# 🛠️ VALID_TIMESHEET_CODES: the codes that have a dedicated grant column on
# the Excel timesheet. Used by the on-screen validation check in Section 17.
VALID_TIMESHEET_CODES = {"NOFA", "WOC", "JJ", "TRICAP", "CARP", "MDHHS", "BEWELL"}
def build_entry_fields(activity, grant):
    """Compose the Activity string, Code, Category, and Description for a log entry."""
    grants = ACTIVITY_STRUCTURE[activity]
    if len(grants) == 1:
        activity_label = activity                      # e.g. "Prevention Team Meeting"
    else:
        activity_label = f"{activity} - {grant}"       # e.g. "Sick - MDHHS", "Training - Botvin"
    return {
        "activity":    activity_label,
        "code":        GRANT_TO_CODE[grant],
        "category":    "Other",
        "description": activity_label,
    }
def generate_time_slots():
    slots = []
    for period in ["AM", "PM"]:
        for hour in range(1, 13):
            for minute in ["00", "15", "30", "45"]:
                slots.append(f"{hour:02d}:{minute} {period}")
    return slots
time_dropdown_options = generate_time_slots()
# =============================================================================
# SECTION 15: DAILY LOG ENTRY FORM
# =============================================================================
# NOTE: This no longer uses st.form. Dependent dropdowns (Grant options change
# based on Activity) don't rerender inside a form, so the widgets live directly
# on the page. Streamlit reruns on each selection, keeping the pair in sync.
st.subheader("⏳ Log Daily Activity")
entry_col1, entry_col2, entry_col3 = st.columns(3)
with entry_col1:
    default_date = TODAY if auto_period_start <= TODAY <= auto_period_end else auto_period_start
    entry_date   = st.date_input(
        "Date Worked", value=default_date,
        min_value=pay_period_start - datetime.timedelta(days=365),
        max_value=pay_period_end   + datetime.timedelta(days=365),
        key="entry_date"
    )
with entry_col2:
    # ── SMART DEFAULTS: preselect the user's last-used times ──
    # Falls back to the original defaults (09:45 AM / 11:30 AM) for new
    # users or if a stored time doesn't match a 15-minute dropdown slot.
    def _slot_index(slot_str, fallback):
        try:
            return time_dropdown_options.index(slot_str)
        except (ValueError, TypeError):
            return fallback
    default_in_idx  = _slot_index(last_time_in_str,  67) if last_time_in_str  else 67
    default_out_idx = _slot_index(last_time_out_str, 74) if last_time_out_str else 74
    time_in_str  = st.selectbox("Time In",  options=time_dropdown_options,
                                index=default_in_idx, key="entry_time_in",
                                help="Defaults to the times from your most recent entry.")
with entry_col3:
    time_out_str = st.selectbox("Time Out", options=time_dropdown_options,
                                index=default_out_idx, key="entry_time_out",
                                help="Defaults to the times from your most recent entry.")
entry_col4, entry_col5 = st.columns(2)
with entry_col4:
    activity_selected = st.selectbox(
        "Activity", sorted(ACTIVITY_STRUCTURE.keys()), key="entry_activity"
    )
with entry_col5:
    grant_options  = ACTIVITY_STRUCTURE[activity_selected]
    grant_selected = st.selectbox(
        "Grant / Program", grant_options, key="entry_grant",
        disabled=(len(grant_options) == 1),
        help="Which grant this time bills to. Locked when the activity only has one funding source."
    )
add_btn = st.button("➕ Save Entry to Log", type="primary")
if add_btn:
    start_time_dt = datetime.datetime.strptime(f"{entry_date} {time_in_str}",  "%Y-%m-%d %I:%M %p")
    end_time_dt   = datetime.datetime.strptime(f"{entry_date} {time_out_str}", "%Y-%m-%d %I:%M %p")
    if end_time_dt <= start_time_dt:
        st.error("Validation Error: 'Time Out' must occur after 'Time In'.")
    else:
        is_dup = check_duplicate(entry_date, time_in_str, time_out_str, instructor_input, all_instructors_df)
        overlaps = [] if is_dup else check_overlap(
            entry_date, start_time_dt, end_time_dt, instructor_input, all_instructors_df
        )
        if is_dup:
            st.markdown(
                '<div class="dup-warning">⚠️ <strong>Duplicate Detected:</strong> An entry with '
                'the same date, time in, and time out already exists in your log. '
                'Submission blocked to prevent double-counting.</div>',
                unsafe_allow_html=True
            )
        elif overlaps:
            st.markdown(
                '<div class="dup-warning">⚠️ <strong>Time Overlap Detected:</strong> This entry '
                f'({time_in_str} – {time_out_str}) overlaps time you already logged on '
                f'{entry_date.strftime("%m/%d/%Y")}: <strong>{" • ".join(overlaps)}</strong>. '
                'Submission blocked to prevent double-counting. Adjust the times so they '
                'don\'t overlap — back-to-back entries (one ending exactly when the next '
                'begins) are fine.</div>',
                unsafe_allow_html=True
            )
        else:
            duration_minutes = int((end_time_dt - start_time_dt).total_seconds() / 60)
            duration_hours   = round(duration_minutes / 60, 2)
            entry_fields     = build_entry_fields(activity_selected, grant_selected)
            form_data = {
                "entry.1205527392": entry_date.strftime("%Y-%m-%d"),
                "entry.1822017875": instructor_input.strip(),
                "entry.1148008178": time_in_str,
                "entry.1036423098": time_out_str,
                "entry.1565734482": entry_fields['activity'],
                "entry.1863736208": entry_fields['code'],
                "entry.835834590":  entry_fields['category'],
                "entry.693720626":  entry_fields['description'],
                "entry.2039394575": duration_minutes,
                "entry.1380701779": duration_hours,
            }

            with st.spinner("Encrypting and saving your entry to the cloud..."):
                try:
                    response = requests.post(LOG_ENTRY_FORM_URL, data=form_data, timeout=10)
                    if response.ok or response.status_code == 200:
                        # Setup the toast message in session state and trigger a refresh
                        st.session_state["toast_message"] = f"Entry saved! {entry_fields['activity']} — {hours_to_hhmm(duration_hours)} logged."
                        fetch_timesheets.clear()
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(f"Submission Error (Code {response.status_code}): Verify Google Form Settings.")
                except Exception as e:
                    st.error(f"Network Connection Error: {e}")
# =============================================================================
# SECTION 16: HISTORY REVIEW & COLOR-CODED TABLE
# =============================================================================
st.markdown("---")
st.subheader("📊 Review Period History")
if total_database_records > 0:
    if not current_period_df.empty:
        st.success(f"🔍 Found {len(current_period_df)} entries for this pay period.")
        col_history_table, col_history_stats = st.columns([3, 1])
        with col_history_table:
            rows_html = ""
            for _, row in current_period_df.iterrows():
                code     = str(row.get('Code', '')).strip()
                hrs_raw  = safe_hours(row.get('Hours', 0))
                hrs_disp = f"{hrs_raw:.2f} ({hours_to_hhmm(hrs_raw)})"
                badge    = code_badge(code)
                rows_html += f"""
                <tr>
                    <td style="padding:6px 10px;border-bottom:1px solid #f0f0f0;">{row.get('Date','')}</td>
                    <td style="padding:6px 10px;border-bottom:1px solid #f0f0f0;">{row.get('Time In','')}</td>
                    <td style="padding:6px 10px;border-bottom:1px solid #f0f0f0;">{row.get('Time Out','')}</td>
                    <td style="padding:6px 10px;border-bottom:1px solid #f0f0f0;max-width:200px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{row.get('Activity','')}</td>
                    <td style="padding:6px 10px;border-bottom:1px solid #f0f0f0;text-align:center;">{badge}</td>
                    <td style="padding:6px 10px;border-bottom:1px solid #f0f0f0;text-align:right;font-weight:600;">{hrs_disp}</td>
                </tr>"""
            table_html = f"""
            <div style="overflow-x:auto;">
            <table style="width:100%;border-collapse:collapse;font-size:14px;font-family:'Calibri',sans-serif;">
                <thead>
                    <tr style="background:#7B2CBF;color:white;">
                        <th style="padding:8px 10px;text-align:left;">Date</th>
                        <th style="padding:8px 10px;text-align:left;">Time In</th>
                        <th style="padding:8px 10px;text-align:left;">Time Out</th>
                        <th style="padding:8px 10px;text-align:left;">Activity</th>
                        <th style="padding:8px 10px;text-align:center;">Code</th>
                        <th style="padding:8px 10px;text-align:right;">Hours</th>
                    </tr>
                </thead>
                <tbody>{rows_html}</tbody>
            </table>
            </div>"""
            st.markdown(table_html, unsafe_allow_html=True)
        with col_history_stats:
            st.metric("Total Hours",   f"{running_hours:.2f} hrs")
            st.metric("HH:MM Format",  hours_to_hhmm(running_hours))
            st.metric("Total Minutes", f"{running_minutes} mins")
            st.metric("Entries",       len(current_period_df))
            if "Code" in current_period_df.columns and "Hours" in current_period_df.columns:
                st.markdown("**Hours by Code:**")
                code_summary = (current_period_df.groupby("Code")["Hours"]
                                .apply(lambda x: pd.to_numeric(x, errors='coerce').fillna(0).sum())
                                .reset_index())
                for _, cs_row in code_summary.iterrows():
                    c   = str(cs_row["Code"])
                    h   = round(float(cs_row["Hours"]), 2)
                    bdg = code_badge(c)
                    st.markdown(f'{bdg} &nbsp; <strong>{h:.2f} hrs</strong> ({hours_to_hhmm(h)})', unsafe_allow_html=True)
        # =============================================================================
        # SECTION 16.5: ENTRY CORRECTION REQUESTS
        # =============================================================================
        # Entries are written to the Google Sheet via a Google Form, which the
        # app cannot edit or delete directly. This workflow lets staff flag a
        # mistake themselves: they pick the entry, describe the fix, and the
        # payroll admin gets an email with everything needed to locate and
        # correct the exact row in the Sheet. The app picks up the change
        # automatically within 60 seconds of the Sheet being edited.
        st.markdown("### ✏️ Fix a Mistake")
        with st.expander("Logged something wrong? Request a correction or removal here."):
            corr_df = current_period_df.reset_index(drop=True)
            entry_labels = [
                f"{row.get('Date','')}  |  {row.get('Time In','')} – {row.get('Time Out','')}  |  "
                f"{row.get('Activity','')}  |  {safe_hours(row.get('Hours',0)):.2f} hrs"
                for _, row in corr_df.iterrows()
            ]
            sel_idx = st.selectbox(
                "Which entry needs fixing?",
                options=range(len(entry_labels)),
                format_func=lambda i: entry_labels[i],
                key="corr_entry_idx"
            )
            sel_row = corr_df.iloc[sel_idx]
            corr_type = st.radio(
                "What needs to happen?",
                ["Change this entry", "Delete this entry entirely"],
                horizontal=True, key="corr_type"
            )
            corr_details = st.text_area(
                "Describe the correction",
                placeholder=("e.g. Time Out should be 03:30 PM, not 05:30 PM — I "
                             "accidentally logged 2 extra hours."),
                key="corr_details",
                help="Be specific: which field is wrong and what it should say instead."
            )
            already_sent_key = f"corr_sent_{sel_idx}_{sel_row.get('Timestamp','')}"
            if st.session_state.get(already_sent_key):
                st.info("✅ A correction request for this entry was already sent this session. "
                        "The entry will remain visible until the admin updates the Google Sheet.")
            elif st.button("📨 Send Correction Request", key="corr_send"):
                if corr_type == "Change this entry" and not corr_details.strip():
                    st.error("Please describe what should change so the admin can apply the fix.")
                else:
                    entry_block = (
                        f"  Instructor      : {instructor_input}\n"
                        f"  Date            : {sel_row.get('Date','')}\n"
                        f"  Time In         : {sel_row.get('Time In','')}\n"
                        f"  Time Out        : {sel_row.get('Time Out','')}\n"
                        f"  Activity        : {sel_row.get('Activity','')}\n"
                        f"  Code            : {sel_row.get('Code','')}\n"
                        f"  Minutes         : {sel_row.get('Minutes','')}\n"
                        f"  Hours           : {sel_row.get('Hours','')}\n"
                        f"  Sheet Timestamp : {sel_row.get('Timestamp','(not available)')}\n"
                    )
                    with st.spinner("Sending your correction request to the payroll admin..."):
                        ok, err = send_correction_email(
                            instructor_input, entry_block, corr_type, corr_details
                        )
                    if ok:
                        st.session_state[already_sent_key] = True
                        st.success(
                            "📬 Correction request sent! The entry stays visible (and still "
                            "counts in your totals) until the admin updates the record — "
                            "usually within one business day. Re-download your Excel files "
                            "after the fix is applied."
                        )
                    else:
                        # SMTP offline or not configured — give the user a copyable
                        # request so the workflow still functions without email.
                        st.warning(
                            "⚠️ The automated email system is currently offline, so the "
                            "request couldn't be sent. Copy the summary below and send it "
                            "to the Payroll Administrator directly:"
                        )
                        st.code(
                            f"CORRECTION REQUEST — {corr_type}\n"
                            f"{entry_block}"
                            f"Requested change: {corr_details.strip() or '(delete this entry)'}",
                            language=None
                        )
        # =============================================================================
        # SECTION 17: PRE-DOWNLOAD VALIDATION CHECK & EXCEL DOWNLOADS
        # =============================================================================
        st.markdown("### 📥 Download Excel Files")
        # ── NEW: ON-SCREEN VALIDATION CHECK ─────────────────────────────────
        # Verifies BEFORE download that every hour in the period is coded to a
        # grant column that exists on the Excel timesheet. If any hours would
        # be missing from the grant columns, the user sees exactly which codes
        # are responsible instead of discovering a mystery mismatch in Excel.
        if "Code" in current_period_df.columns and "Hours" in current_period_df.columns:
            _hours_num  = pd.to_numeric(current_period_df["Hours"], errors="coerce").fillna(0.0)
            _codes_up   = current_period_df["Code"].astype(str).str.strip().str.upper()
            mapped_total   = round(float(_hours_num[_codes_up.isin(VALID_TIMESHEET_CODES)].sum()), 2)
            overall_total  = round(float(_hours_num.sum()), 2)
            check_diff     = round(overall_total - mapped_total, 2)
            if abs(check_diff) >= 0.01:
                bad_codes = sorted(set(_codes_up[~_codes_up.isin(VALID_TIMESHEET_CODES)].tolist()))
                st.error(
                    f"⚠️ **Validation Check Failed:** {check_diff:.2f} hrs this period are logged "
                    f"under unrecognized grant code(s) — {', '.join(bad_codes) if bad_codes else 'unknown'} — "
                    f"and will NOT appear in any grant column on the Excel timesheet. "
                    f"The file will include a warning row and its built-in Balance Check will flag the gap. "
                    f"Fix these entries in the Google Sheet to fully balance."
                )
            else:
                st.success(
                    f"✅ **Validation Check Passed:** Grant column totals ({mapped_total:.2f} hrs) "
                    f"match the period total ({overall_total:.2f} hrs / {hours_to_hhmm(overall_total)}). "
                    f"The Excel file also contains a live Column Subtotals row and Balance Check cell."
                )
        col_dl1, col_dl2 = st.columns(2)
        safe_name = instructor_input.replace(" ", "_")
        # Serialise period data for cache key (orient="records" handles dataframe rows cleanly)
        period_json = current_period_df.to_json(orient="records", date_format="iso")
        today_str   = TODAY.strftime("%m/%d/%Y")
        with col_dl1:
            ts_bytes = build_timesheet_bytes(instructor_input, pay_period_start, pay_period_end, period_json, today_str)
            st.download_button(
                label="📥 Download Timesheet (.xlsx)",
                data=ts_bytes,
                file_name=f"{safe_name}_Official_Timesheet_{pay_period_start}_to_{pay_period_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col_dl2:
            add_bytes = build_additional_bytes(instructor_input, pay_period_start, pay_period_end, period_json)
            st.download_button(
                label="📥 Download Additional Hours Report (.xlsx)",
                data=add_bytes,
                file_name=f"{safe_name}_Additional_Hours_{pay_period_start}_to_{pay_period_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning(
            f"ℹ️ **{total_database_records} total entries** found for your profile, but "
            f"**0 entries** fall within {pay_period_start.strftime('%m/%d/%Y')} – "
            f"{pay_period_end.strftime('%m/%d/%Y')}."
        )
        st.info("💡 Adjust the Start/End Date inputs above to reveal your logged entries.")
# =============================================================================
# SECTION 18: ADMIN PANEL  (only visible when logged in with master password)
# =============================================================================
if is_admin and not all_instructors_df.empty:
    st.markdown("---")
    st.markdown("## 🛡️ Admin Dashboard")
    st.caption("Visible only to admin accounts.")
    period_all = all_instructors_df[
        (all_instructors_df["ParsedDate"] >= pay_period_start) &
        (all_instructors_df["ParsedDate"] <= pay_period_end)
    ].copy()
    if not period_all.empty and "Instructor Name" in period_all.columns:
        st.markdown("### 👥 Team Hours — Current Period")
        registered_names = {
            str(n).strip()
            for n in (account_registry["Instructor Name"].tolist()
                      if not account_registry.empty else [])
            if str(n).strip() and str(n).strip().lower() != "nan"
        }
        summary = (
            period_all.groupby("Instructor Name")
            .agg(
                Hours=("Hours", lambda x: pd.to_numeric(x, errors='coerce').fillna(0).sum()),
                Entries=("Hours", "count")
            )
            .reset_index()
            .sort_values("Hours", ascending=False)
        )
        summary["HH:MM"]      = summary["Hours"].apply(hours_to_hhmm)
        summary["Hours"]      = summary["Hours"].apply(lambda x: f"{x:.2f}")
        summary["⚠️ No Log"]  = summary["Instructor Name"].apply(lambda n: "✅" if n in registered_names else "—")
        logged_names    = {
            str(n).strip()
            for n in period_all["Instructor Name"].tolist()
            if str(n).strip() and str(n).strip().lower() != "nan"
        }
        missing_names   = registered_names - logged_names
        if missing_names:
            st.warning(
                f"⚠️ **{len(missing_names)} instructor(s) have not logged any time this period:** "
                + ", ".join(sorted(missing_names))
            )
        st.dataframe(summary, use_container_width=True, hide_index=True)
        st.markdown("### 📊 Team Hours by Code — Current Period")
        if "Code" in period_all.columns:
            code_totals = (
                period_all.groupby("Code")["Hours"]
                .apply(lambda x: pd.to_numeric(x, errors='coerce').fillna(0).sum())
                .reset_index()
                .sort_values("Hours", ascending=False)
            )
            cols_admin = st.columns(len(code_totals))
            for i, (_, cr) in enumerate(code_totals.iterrows()):
                c    = str(cr["Code"])
                h    = round(float(cr["Hours"]), 2)
                bdg  = code_badge(c)
                with cols_admin[i]:
                    st.markdown(
                        f'<div style="text-align:center;padding:12px;background:#f8f5ff;'
                        f'border-radius:10px;">{bdg}<br>'
                        f'<span style="font-size:22px;font-weight:800;">{h:.2f}</span><br>'
                        f'<span style="font-size:12px;color:#64748b;">hrs ({hours_to_hhmm(h)})</span>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
        # ── PER-INSTRUCTOR DETAIL DRILL-DOWN ──
        # Pick one person and see every day, activity, code, and time they
        # logged this period — the same detail an instructor sees for
        # themselves, but for any team member.
        st.markdown("### 🔎 Individual Instructor Detail")
        instructor_choices = sorted(
            {str(n).strip() for n in period_all["Instructor Name"].tolist()
             if str(n).strip() and str(n).strip().lower() != "nan"}
        )
        if instructor_choices:
            picked = st.selectbox(
                "Select an instructor to see their day-by-day detail:",
                options=instructor_choices,
                key="admin_detail_pick"
            )
            person_df = period_all[
                period_all["Instructor Name"].astype(str).str.strip() == picked
            ].copy()
            if "ParsedDate" in person_df.columns:
                person_df = person_df.sort_values("ParsedDate")
            elif "Date" in person_df.columns:
                person_df = person_df.sort_values("Date")
            p_hours   = round(float(pd.to_numeric(person_df.get("Hours", 0), errors="coerce").fillna(0).sum()), 2)
            p_minutes = int(pd.to_numeric(person_df.get("Minutes", 0), errors="coerce").fillna(0).sum())
            p_entries = len(person_df)
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total Hours",   f"{p_hours:.2f}")
            m2.metric("HH:MM",         hours_to_hhmm(p_hours))
            m3.metric("Total Minutes", f"{p_minutes}")
            m4.metric("Entries",       p_entries)
            # Day-by-day detail table
            detail_cols = [c for c in ['Date','Time In','Time Out','Activity','Code','Hours','Minutes']
                           if c in person_df.columns]
            st.dataframe(
                person_df[detail_cols],
                use_container_width=True, hide_index=True
            )
            # Per-grant breakdown for the selected person
            if "Code" in person_df.columns and "Hours" in person_df.columns:
                st.markdown(f"**{picked} — hours by grant code:**")
                pcode = (person_df.groupby("Code")["Hours"]
                         .apply(lambda x: round(float(pd.to_numeric(x, errors="coerce").fillna(0).sum()), 2))
                         .reset_index()
                         .sort_values("Hours", ascending=False))
                for _, prow in pcode.iterrows():
                    cc = str(prow["Code"]); hh = float(prow["Hours"])
                    st.markdown(
                        f'{code_badge(cc)} &nbsp; <strong>{hh:.2f} hrs</strong> ({hours_to_hhmm(hh)})',
                        unsafe_allow_html=True
                    )
            # Let admin download this one person's timesheet + additional report
            _pj = person_df.to_json(orient="records", date_format="iso")
            _safe = picked.replace(" ", "_")
            dcol1, dcol2 = st.columns(2)
            with dcol1:
                st.download_button(
                    "📥 Download This Instructor's Timesheet (.xlsx)",
                    data=build_timesheet_bytes(picked, pay_period_start, pay_period_end,
                                               _pj, TODAY.strftime("%m/%d/%Y")),
                    file_name=f"{_safe}_Official_Timesheet_{pay_period_start}_to_{pay_period_end}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="admin_dl_ts"
                )
            with dcol2:
                st.download_button(
                    "📥 Download Additional Hours Report (.xlsx)",
                    data=build_additional_bytes(picked, pay_period_start, pay_period_end, _pj),
                    file_name=f"{_safe}_Additional_Hours_{pay_period_start}_to_{pay_period_end}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="admin_dl_add"
                )
        with st.expander("📋 View Full Team Log for This Period"):
            cols_to_show = [c for c in ['Date','Instructor Name','Activity','Code','Hours','Minutes'] if c in period_all.columns]
            st.dataframe(period_all[cols_to_show].sort_values(["Instructor Name","Date"]), use_container_width=True, hide_index=True)

    else:
        st.info("ℹ️ No hours have been logged by any team members for this pay period yet.")
